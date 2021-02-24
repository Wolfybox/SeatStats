import argparse
import re

from openpyxl.styles import Alignment
from openpyxl import load_workbook

import xlrd


class SeatHelper:
    def __init__(self, sign_info_dir, stu_info_dir):
        stu_data = xlrd.open_workbook(stu_info_dir)
        stu_sheet = stu_data.sheet_by_index(0)
        print(f'Total Students: {stu_sheet.nrows}')
        target_stu_list = []
        for row_index in range(stu_sheet.nrows):
            row_ele_list = stu_sheet.row_values(row_index)
            row_id, stu_id, stu_name, stu_acade, _, stu_class, _, _, ta_name = row_ele_list
            target_stu_list.append(stu_name)

        sign_data = xlrd.open_workbook(sign_info_dir)
        sign_tables = sign_data.sheet_by_name('签到查询')
        row_num, col_num = sign_tables.nrows, sign_tables.ncols
        print(f'Total Row:{row_num} Col:{col_num}')
        self.stu_id, self.stu_name, self.stu_seat, self.stu_time = [], [], [], []
        # self.stu_info = {}
        for row_index in range(1, row_num):
            row_ele_list = sign_tables.row_values(row_index)
            print(f'Row {row_index} : {row_ele_list}')
            row_id, row_name, row_acade, row_seat, row_time = row_ele_list
            # self.stu_info[row_name] = {
            #     'id': row_id,
            #     'acade': row_acade,
            #     'seat': row_seat,
            #     'time': row_time
            # }
            if row_name not in self.stu_name and row_name in target_stu_list:
                self.stu_id.append(row_id)
                self.stu_name.append(row_name)
                self.stu_seat.append(row_seat)
                self.stu_time.append(row_time)


def generate_seat_map(template_dir, seat_helper, save_dir):
    ori_data = xlrd.open_workbook(template_dir)
    ori_tables = ori_data.sheet_by_name('模板')
    dst_wb = load_workbook(template_dir)
    dst_wb._active_sheet_index = 1
    seat_tables = dst_wb.active
    nrows, ncols = ori_tables.nrows, ori_tables.ncols

    # get row cols indices
    row_indices, col_indices = [], []
    row_pattern, col_pattern = re.compile(r'第.{1,3}排'), re.compile(r'第.{1,3}列')
    for row_i in range(nrows):
        for col_i in range(ncols):
            cell = ori_tables.cell_value(row_i, col_i)
            if re.match(row_pattern, cell) is not None:
                row_indices.append(row_i)
            if re.match(col_pattern, cell) is not None:
                col_indices.append(col_i)

    for i in range(len(seat_helper.stu_name)):
        cname, cseat = seat_helper.stu_name[i], seat_helper.stu_seat[i]
        seat_elements = cseat.split('-')
        classroom, room_row, room_col = seat_elements[0], int(seat_elements[1]), int(seat_elements[2])
        seat_tables.cell(row=row_indices[room_row - 1] + 1, column=col_indices[room_col - 1] + 1,
                         value=f'{cseat}\n{cname}').alignment = Alignment(wrap_text=True, horizontal='center',
                                                                          vertical='center')
    dst_wb.save(save_dir)


if __name__ == '__main__':
    argparser = argparse.ArgumentParser()
    argparser.add_argument('--sign-info', type=str, default='签到查询.xls', help='座位系统导出的签到表')
    argparser.add_argument('--stu-info', type=str, default=r'大学计算机B班（自动化5-8班）.xlsx', help='学生姓名表')
    argparser.add_argument('--save-dir', type=str, default='周三56节T3201座位表.xlsx', help='输出的座位表')
    argparser.add_argument('--template-dir', type=str, default='附件3：固定座位图模板.xlsx', help='座位表模板')
    config = argparser.parse_args()
    seat_helper = SeatHelper(sign_info_dir=config.sign_info, stu_info_dir=config.stu_info)
    generate_seat_map(template_dir=config.template_dir, seat_helper=seat_helper,
                      save_dir=config.save_dir)
